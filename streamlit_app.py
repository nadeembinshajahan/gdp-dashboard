import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import base64

# Title and description
st.title("Attendance Report Generator")
st.write("Upload a report.csv file to generate an attendance report with employee information.")

# Function to clean report data
def clean_report_data(report_df):
    """
    Clean the report CSV data by filtering out invalid rows and fixing formats.
    """
    # Filter only rows with valid Person ID and Date
    valid_rows = report_df[
        (report_df['Person ID'].notna()) & 
        (report_df['Person ID'].str.strip() != '') & 
        (~report_df['Person ID'].astype(str).str.contains('Check-In Time')) &
        (report_df['Date'].notna()) & 
        (report_df['Date'].str.strip() != '')
    ]
    
    # Clean Person IDs by removing leading apostrophes
    valid_rows.loc[:, 'Person ID'] = valid_rows['Person ID'].astype(str).str.replace("^'", "", regex=True)
    
    return valid_rows

# Function to get day of week
def get_day_of_week(date_str):
    """
    Get the day of the week from a date string.
    """
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%A')
    except:
        return ""

# Function to calculate hours
def calculate_hours(check_in, check_out):
    """
    Calculate the total hours between check-in and check-out times.
    Returns hours as a float and the duration as HH:MM:SS.
    """
    if not isinstance(check_in, str) or not isinstance(check_out, str):
        return 0.0, "00:00:00"
    
    if check_in == '-' or check_out == '-' or not check_in or not check_out:
        return 0.0, "00:00:00"
    
    try:
        # Parse hours, minutes, seconds
        in_parts = check_in.split(':')
        out_parts = check_out.split(':')
        
        if len(in_parts) < 2 or len(out_parts) < 2:
            return 0.0, "00:00:00"
        
        # Handle the case when seconds might not be present
        in_hours, in_minutes = int(in_parts[0]), int(in_parts[1])
        in_seconds = int(in_parts[2]) if len(in_parts) > 2 else 0
        
        out_hours, out_minutes = int(out_parts[0]), int(out_parts[1])
        out_seconds = int(out_parts[2]) if len(out_parts) > 2 else 0
        
        # Calculate total seconds
        total_in_seconds = in_hours * 3600 + in_minutes * 60 + in_seconds
        total_out_seconds = out_hours * 3600 + out_minutes * 60 + out_seconds
        
        # Calculate difference in seconds
        diff_seconds = total_out_seconds - total_in_seconds
        
        # Format as HH:MM:SS
        hours = diff_seconds // 3600
        minutes = (diff_seconds % 3600) // 60
        seconds = diff_seconds % 60
        
        time_format = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
        # Also return the hours as float for calculations
        return round(diff_seconds / 3600, 2), time_format
    except:
        return 0.0, "00:00:00"

# Function to calculate total minutes
def calculate_total_minutes(hours):
    """
    Convert hours to minutes.
    """
    return round(hours * 60)

# Function to calculate extra minutes
def calculate_extra_minutes(total_minutes, standard_minutes=540):
    """
    Calculate extra minutes beyond standard working time (default: 9 hours = 540 minutes).
    """
    if total_minutes <= standard_minutes:
        return 0
    return total_minutes - standard_minutes

# Function to create a download link for the DataFrame
def get_excel_download_link(df, filename="attendance_report.xlsx"):
    """
    Generate a download link for an Excel file from a DataFrame.
    """
    # Create a buffer to write the Excel file
    output = io.BytesIO()
    
    # Use ExcelWriter with formatting options
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance Report')
        
        # Get the worksheet
        worksheet = writer.sheets['Attendance Report']
        
        # Import styles
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Define header style
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Define status colors
        normal_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        absence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        
        # Define employee alternating colors (pastel shades)
        employee_colors = [
            "FFFFFF",  # White (default)
            "F2F2F2",  # Light gray
            "EBF1DE",  # Light mint
            "E6E0EC",  # Light lavender
            "DAEEF3",  # Light blue
            "FFEBCC",  # Light orange
            "F2DCDB",  # Light pink
            "EBF1DE",  # Light green
            "E6E0EC"   # Light purple
        ]
        
        # Apply styles to headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
        
        # Apply conditional formatting and employee row coloring
        current_employee = None
        color_index = 0
        attendance_col_index = df.columns.get_loc("ATTENDANCE STATUS") + 1
        
        for row_idx, row in enumerate(df.itertuples(), 2):  # Start from 2 to account for header
            # Check if employee changed
            if current_employee != row.NEW_EMPLOYEE_NO_:
                current_employee = row.NEW_EMPLOYEE_NO_
                color_index = (color_index + 1) % len(employee_colors)
            
            # Apply row background color based on employee
            row_fill = PatternFill(start_color=employee_colors[color_index], 
                                  end_color=employee_colors[color_index], 
                                  fill_type="solid")
            
            # Apply row color
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
            
            # Highlight attendance status
            status_cell = worksheet.cell(row=row_idx, column=attendance_col_index)
            status_value = str(status_cell.value).strip().lower()
            
            if status_value == "normal":
                status_cell.fill = normal_fill
                status_cell.font = Font(color="006100", bold=True)  # Dark green text
            elif status_value == "absence":
                status_cell.fill = absence_fill
                status_cell.font = Font(color="9C0006", bold=True)  # Dark red text
        
        # Add borders to all cells
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Adjust column widths to fit content
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Find the maximum length in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set the column width (with a minimum and maximum)
            adjusted_width = min(max(max_length + 2, 10), 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header to ensure visibility with wrap_text
            worksheet.row_dimensions[1].height = 40
    
    # Get the value of the buffer
    excel_data = output.getvalue()
    
    # Create a download link
    b64 = base64.b64encode(excel_data).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}" class="btn btn-primary">Download Excel Report</a>'
    return href

# Create an embedded database of employee information (extracted from database.csv)
# This is hardcoded to avoid needing to upload database.csv every time
def get_employee_database():
    """
    Returns a DataFrame with the employee database information.
    """
    # Database is embedded directly in the code
    data = """Person ID,"Employee ID
","Employee Name
","Department
",,,,,,du
1000,1000,THASLEEM VP,COO,,,,,,
1001,1001,SHABIL CHAKKOLAYIL,CSO,,,,,,
S003,TTQA003,Abdul Shifas,Pre-Sales,,,,,,
S004,TTQA004,Shameer Chundekattil Abu,Technical,,,,,,
S006,TTQA006,Shyamraj Sasidharan,Pre-Sales,,,,,,
S007,TTQA007,Jobin Therath Joseph,Technical,,,,,,
S012,TTQA012,Maneesh Mani Koroththazha,Procurement,,,,,,
S021,TTQA021,Kishanthan Thiyagarajan,Technical,,,,,,
S027,TTQA027,Anoop Aniyan,Finance and Accounts,,,,,,
,TTQA040,Surendran Vinayagam Surendran,Technical,,,,,,
S042,TTQA042,Shuhaib Muhammed Kutty,Human Resources and Admin.,,,,,,
S088,TTQA088,Ratheesh Ittikkatt Velayudhan,Procurement,,,,,,
S101,TTQA101,Jishnu Neriathraparambil Shibu,Technical,,,,,,
S118,TTQA118,Mohammed Nasfer Valoth,Renewals - Licensing,,,,,,
S124,TTQA124,Shamsudheen Kalli Kuniyil,Human Resources and Admin.,,,,,,
S146,TTQA146,Muhammad Farash Kunhippurayil,Technical,,,,,,
S151,TTQA151,Mohammed Ali Chelathoden Latif,Pre-Sales,,,,,,
S160,TTQA160,Asharaf Naranath,Finance and Accounts,,,,,,
S171,TTQA171,Nabil Jamal Vettikkal,Sales,,,,,,
S173,TTQA173,Sreelaskshmi Pottekkatt Sudhi,Marketing,,,,,,
S174,TTQA174,Irfan Kottukkal Ismail,Support,,,,,,
TTQ200,TTQA200,Basheer Andanakathu Moidunni,Human Resources and Admin.,,,,,,
S102,TTQA201,Sumesh K Sami,Design & Lighting,,,,,,
S203,TTQA203,Muhammad Musthafa,Pre-Sales,,,,,,
S206,TTQA207,Niyas Vayalil Muhammad Ismail,Human Resources and Admin.,,,,,,
S221,TTQA221,Muhammad Hassan Malik,Technical,,,,,,
TTQ219,TTQA224,Kirusanth Velayutham Anandarasa,Technical,,,,,,
S225,TTQA225,Varun Puthan Purayil,Technical,,,,,,
TTQ226,TTQA226,Yadhu Krishnan,Design & Lighting,,,,,,
S214,TTQA244,Pranav Ponath Uthaman,Design & Lighting,,,,,,
QA248,TTQA248,Muhammed Ali Jawahar Shajahan Naseela,Support,,,,,,
QA250,TTQA250,Muhammad Abid Nasir,Renewals - Licensing,,,,,,
S251,TTQA251,Shabin Ashif N. Parambil,Human Resources and Admin.,,,,,,
TTQ252,TTQA252,Nidhisha Chandran,Finance and Accounts,,,,,,
TTQ253,TTQA253,Mohamad Kawwam,Marketing,,,,,,
TTQ254,TTQA254,Rafiyath Rafeeque Moidu,Pre-Sales,,,,,,
TTQ255,TTQA255,Mudhasar Olavakkott S,Support,,,,,,
TTQ263,TTQA263,Rocky Rajan Varughese,Procurement,,,,,,
TTQ219,TTQA272,Nawaz Shereef Kareem,Procurement,,,,,,
TTQ274,TTQA274,Burhan Mohammad Walid Alhaffar,Technical,,,,,,
TTQ275,TTQA275,Balagopal Venugopal Kanika,Procurement,,,,,,
TTQ281,TTQA281,Allwyn Solomon Rajit Singh,Support,,,,,,
TTQ285,TTQA285,Jino Jose,Design & Lighting,,,,,,
TTQ286,TTQA286,Anish Kumar Murugan,Technical,,,,,,
TTQ288,TTQA288,Zeeshan Ul Haq,Pre-Sales,,,,,,
TTQ298,TTQA298,Afsal Veluthandi Madathummal,Support,,,,,,
TTQ301,TTQA301,Aafaq Burhanuddin Tisekar,Human Resources and Admin.,,,,,,
TTQ302,TTQA302,Anshad Kelathum Parambil,Finance and Accounts,,,,,,
TTQA303,TTQA303,Le Mogene De la Paz Tucker,Human Resources and Admin.,,,,,,
TTQA304,TTQA304,Nihad El Harch,Sales,,,,,,
TTQA308,TTQA308,Mohammad Mamoon Mohammad Haroon,Pre-Sales,,,,,,
TTQA309,TTQA309,Tiago Andre Pereira Reis,Sales,,,,,,
TTQA310,TTQA310,Akhil Thoppil Aliyar,Sales,,,,,,
TTQA311,TTQA311,Mohammed Muzzammil Pradanakkaar,Sales,,,,,,
TTQA312,TTQA312,Abdallah Mohammed Abdelhakem Elkady,Sales,,,,,,
TTQA313,TTQA313,Mohammed Osman,Pre-Sales,,,,,,
,TTQA314,Kamal Navas,Pre-Sales,,,,,,
,TTQA316,Suhail Thayyil,HRA,,,,,,
"""
    # Convert the string data to a DataFrame and clean up column names
    from io import StringIO
    df = pd.read_csv(StringIO(data))
    
    # Clean up column names - removing newlines
    df.columns = [col.replace('\n', '').strip() for col in df.columns]
    
    return df

# File uploader for report.csv
uploaded_file = st.file_uploader("Upload report.csv", type="csv")

if uploaded_file is not None:
    # Process the uploaded file
    try:
        report_df = pd.read_csv(uploaded_file)
        st.success("File uploaded successfully!")
        
        # Display the first few rows of the report
        with st.expander("Preview of uploaded data"):
            st.dataframe(report_df.head())
        
        # Get the embedded database
        db_df = get_employee_database()
        
        # Clean the data
        report_df = clean_report_data(report_df)
        
        # Create a mapping from Person ID to employee details
        employee_mapping = {}
        for _, row in db_df.iterrows():
            if pd.notna(row['Person ID']) and str(row['Person ID']).strip():
                employee_mapping[str(row['Person ID']).strip()] = {
                    'employee_id': row.get('Employee ID', ''),
                    'employee_name': row.get('Employee Name', ''),
                    'department': row.get('Department', '')
                }
        
        # Create the output data
        output_data = []
        
        for _, row in report_df.iterrows():
            person_id = str(row['Person ID']).strip()
            
            # Skip if the person ID is not in the database mapping
            if person_id not in employee_mapping:
                continue
            
            # Get employee details
            employee_details = employee_mapping[person_id]
            
            # Calculate time values
            check_in = row.get('Check-In', '-')
            check_out = row.get('Check-out', '-')
            
            hours, hours_formatted = calculate_hours(check_in, check_out)
            total_minutes = calculate_total_minutes(hours)
            extra_minutes = calculate_extra_minutes(total_minutes)
            
            # Prepare the output row
            output_row = {
                'DATE': row.get('Date', ''),
                'DAY': get_day_of_week(row.get('Date', '')),
                'NEW EMPLOYEE NO.': employee_details['employee_id'],
                'EMPLOYEE NAME': employee_details['employee_name'],
                'DEPARTMENT NAME': employee_details['department'],
                'TIME IN': '' if check_in == '-' else check_in,
                'TIME OUT': '' if check_out == '-' else check_out,
                'TOTAL HOURS': hours_formatted,
                'ATTENDANCE STATUS': row.get('Attendance Status', ''),
                'EXTRA MINUTES': str(extra_minutes),
                'NEW INTIME': '',
                'NEW OUTTIME': '',
                'NEW TOTAL HOURS': '',
                'OT/UT': '',
                'REMARKS': '',
                'REASON': '',
                'MANAGER\'S COMMENTS': ''
            }
            
            output_data.append(output_row)
        
        # Create DataFrame from output data
        output_df = pd.DataFrame(output_data)
        
        # Sort the DataFrame by NEW EMPLOYEE NO and DATE (with latest dates last)
        output_df['DATE'] = pd.to_datetime(output_df['DATE'], errors='coerce')
        output_df = output_df.sort_values(['NEW EMPLOYEE NO.', 'DATE'])
        output_df['DATE'] = output_df['DATE'].dt.strftime('%Y-%m-%d')
        
        # Display statistics
        st.subheader("Report Statistics")
        col1, col2, col3 = st.columns(3)
        col1.metric("Total Records", len(report_df))
        col2.metric("Matched Records", len(output_data))
        col3.metric("Employees Covered", len(output_df['NEW EMPLOYEE NO.'].unique()))
        
        # Display a preview of the output
        st.subheader("Preview of Generated Report")
        st.dataframe(output_df.head(10))
        
        # Provide a download link
        st.subheader("Download Report")
        st.markdown(get_excel_download_link(output_df), unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"Error processing file: {e}")
else:
    st.info("Please upload a report.csv file to generate the attendance report.")

# Add a footer with instructions
st.markdown("---")
st.markdown("""
### Instructions:
1. Upload your report.csv file using the file uploader above
2. The app will automatically process the file and generate an attendance report
3. Download the Excel file using the link provided
4. The employee database is built into the app, so you don't need to upload database.csv

### Note:
- The report includes calculations for total hours (in HH:MM:SS format) and extra minutes
- All entries for the same employee are grouped together and sorted by date
""")

# Sidebar with additional information
with st.sidebar:
    st.header("About this App")
    st.write("""
    This app generates an attendance report by processing raw attendance data and matching it with employee information.
    
    The employee database is embedded in the app, so you only need to upload the attendance report file (report.csv).
    
    The output includes the following columns:
    - DATE
    - DAY
    - NEW EMPLOYEE NO.
    - EMPLOYEE NAME
    - DEPARTMENT NAME
    - TIME IN
    - TIME OUT
    - TOTAL HOURS (in HH:MM:SS format)
    - ATTENDANCE STATUS
    - EXTRA MINUTES
    - NEW INTIME (empty)
    - NEW OUTTIME (empty)
    - NEW TOTAL HOURS (empty)
    - OT/UT (empty)
    - REMARKS (empty)
    - REASON (empty)
    - MANAGER'S COMMENTS (empty)
    """)
